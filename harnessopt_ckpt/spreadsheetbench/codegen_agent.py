"""Codegen agent for SpreadsheetBench — no tool-call, pure code generation.

Two modes:
  - **single**: One LLM call → extract ```python``` block → done.
  - **multi**: Up to max_turns LLM calls; after each, execute code and
    feed errors back for correction.

This matches the official SpreadsheetBench evaluation setting (LLM generates
a Python code block, no function-calling / tool-use).
"""
from __future__ import annotations

import json
import os
import random
import signal
import time

import openpyxl


# ── Timeout helper ──────────────────────────────────────────────────────────

class TaskTimeout(Exception):
    """Raised when a task exceeds its time budget."""


def _timeout_handler(signum, frame):
    raise TaskTimeout("Task timed out")

from skillopt.model.azure_openai import (
    get_reasoning_effort,
    get_target_client,
    _needs_responses_api,
    tracker,
)
from skillopt.model import get_codex_exec_config, get_target_backend, is_target_exec_backend
from skillopt.model.codex_harness import prepare_workspace, render_skill_md, run_target_exec
from skillopt.prompts import load_prompt
from .executor import run_generated_code
from .evaluator import evaluate


# ── Eval feedback helper (no golden value leakage) ─────────────────────────

def _build_eval_feedback(verify_report: str) -> str:
    """Build Target feedback from a verify report, hiding expected values.

    The verify report contains lines like:
        Sheet1!D2: got=None, expected=0 ✗
        Sheet1!D10: got=None, expected=None ✓

    We strip the ``expected=...`` part so the Target sees only its own
    output and whether each cell is correct or wrong.
    """
    import re
    lines = ["Your code executed successfully but produced incorrect results.",
             "The following cells have wrong values:"]
    for raw_line in verify_report.splitlines():
        raw_line = raw_line.strip()
        if not raw_line:
            continue
        # Match enrichment lines like "  Sheet1!D2: got=None, expected=0 ✗"
        m = re.match(
            r"(\S+!?\w+):\s*got=(.+?),\s*expected=.+?\s*(✓|✗)$",
            raw_line,
        )
        if m:
            cell, got_val, mark = m.groups()
            if mark == "✗":
                lines.append(f"  {cell}: your output = {got_val} (WRONG)")
            else:
                lines.append(f"  {cell}: correct ✓")
    lines.append(
        "\nPlease analyze the spreadsheet data more carefully and fix the code. "
        "Return a complete corrected Python script inside a ```python``` block."
    )
    return "\n".join(lines)


# ── Workbook preview (same as official prompt.py) ────────────────────────────
#
# The `harness_nano` bundle (SPREADSHEETBENCH_HARNESS_NANO=1) enables
# round-0-best behavior:  bigger preview (5x20 -> 15x30) + gold-free
# self-introspect retry in run_multi.  The bundle was tuned for
# gpt-5.4-nano (test_hard 0.7402, val_hard 0.7436 vs baseline 0.6410).
# It is DEFAULT OFF so the harness ships as pristine baseline; enable
# explicitly with SPREADSHEETBENCH_HARNESS_NANO=1.
#
# Individual overrides (SPREADSHEETBENCH_PREVIEW_ROWS / _COLS) always win
# and take effect regardless of the master switch, so you can tune the
# preview budget without pulling in introspect.
_HARNESS_NANO = os.environ.get("SPREADSHEETBENCH_HARNESS_NANO", "0") == "1"
_PREVIEW_ROWS_DEFAULT = int(os.environ.get(
    "SPREADSHEETBENCH_PREVIEW_ROWS", "15" if _HARNESS_NANO else "5"
))
_PREVIEW_COLS_DEFAULT = int(os.environ.get(
    "SPREADSHEETBENCH_PREVIEW_COLS", "30" if _HARNESS_NANO else "20"
))


# ── Reasoning-effort timeout fallback ────────────────────────────────────────
#
# Some gpt-5.x reasoning tasks silently run away server-side and never emit
# an assistant message before the client-side task_timeout (600 s) fires,
# giving `turns=0 dt=600s` in the log. Setting SPREADSHEETBENCH_TIMEOUT_FALLBACK=1
# tells run_multi that if the first LLM call fails (timeout / retry exhaustion),
# it should retry once with a lower reasoning effort (default `low`) so the
# task can at least produce a solution attempt instead of returning zero turns.
#
# Trade-offs:
#   * When active, the first phase uses `retries=2` (instead of 5) so we cap
#     wasted time at ~2 * llm_timeout (~240 s) and reserve budget for the
#     fallback attempt. Transient 5xx errors thus retry fewer times too.
#   * Only applies to the Chat / Responses backend (not the codex/exec backend).
#   * No-op when the configured effort is already "low" or "minimal", or when
#     get_reasoning_effort() returns None (non-reasoning models).
_TIMEOUT_FALLBACK = os.environ.get("SPREADSHEETBENCH_TIMEOUT_FALLBACK", "0") == "1"
_FALLBACK_EFFORT = os.environ.get("SPREADSHEETBENCH_TIMEOUT_FALLBACK_EFFORT", "low")


def _preview_workbook(path: str, max_rows: int | None = None, max_cols: int | None = None) -> str:
    """Generate a text preview of the first few rows of each sheet."""
    if max_rows is None:
        max_rows = _PREVIEW_ROWS_DEFAULT
    if max_cols is None:
        max_cols = _PREVIEW_COLS_DEFAULT
    wb = openpyxl.load_workbook(path, data_only=False)
    chunks: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        chunks.append(
            f"## Sheet: {sheet_name}  "
            f"(dim={ws.dimensions}, max_row={ws.max_row}, max_col={ws.max_column})"
        )
        for row in ws.iter_rows(
            min_row=1,
            max_row=min(ws.max_row, max_rows),
            max_col=min(ws.max_column, max_cols),
            values_only=False,
        ):
            cells = []
            for cell in row:
                v = cell.value
                if v is None:
                    cells.append(f"{cell.coordinate}=")
                else:
                    s = str(v)
                    if len(s) > 40:
                        s = s[:37] + "..."
                    cells.append(f"{cell.coordinate}={s}")
            chunks.append(" | ".join(cells))
        if ws.max_row > max_rows:
            chunks.append(f"... ({ws.max_row - max_rows} more rows)")
        chunks.append("")
    wb.close()
    return "\n".join(chunks)


# ── Self-introspect (gold-free post-exec sanity check) ──────────────────────
#
# round-0 harness-opt: when run_multi succeeds at exec but has no gold to
# compare against (gold_path="" in eval/val/test, by design), the loop
# previously broke at turn 1 and missed the chance to recover from
# structurally-broken outputs. This helper inspects the answer-position
# cells in the agent's *own* output and flags two structural bugs that
# need NO gold to detect:
#   1. target cell is None (cluster C1 "no-computation", 5/29 train failures)
#   2. target cell holds a literal "=..." formula string (skill already says
#      not to, but the agent still does it; grader reads cached value = None
#      and fails)
# When flagged, run_multi re-prompts the agent with a gold-free feedback
# message (the actual current values, NOT the expected ones).
#
# Gated by the `harness_nano` master switch: introspect only fires when
# SPREADSHEETBENCH_HARNESS_NANO=1 AND SPREADSHEETBENCH_SELF_INTROSPECT != 0.
# Since HARNESS_NANO defaults off, the retry loop is disabled by default.
_SELF_INTROSPECT_ENABLED = (
    _HARNESS_NANO
    and os.environ.get("SPREADSHEETBENCH_SELF_INTROSPECT", "1") == "1"
)


def _self_introspect(output_path: str, answer_position: str) -> str | None:
    """Return ``None`` if the agent's output looks structurally fine at
    ``answer_position``; otherwise return a feedback string for re-prompting.

    Never reveals the gold value — only what the agent's own cells contain.
    """
    if not _SELF_INTROSPECT_ENABLED:
        return None
    if not output_path or not answer_position:
        return None
    if not os.path.exists(output_path):
        return None
    try:
        from .evaluator import _generate_cell_names
        # data_only=False so we can see literal formula strings the agent left in;
        # openpyxl with data_only=True would coerce those to None and we'd miss them.
        wb = openpyxl.load_workbook(output_path, data_only=False)
    except Exception:
        return None

    # Note: we deliberately do NOT flag None/empty cells as suspicious. Gold
    # answer-position ranges commonly contain legitimately-empty cells (the
    # instruction only fills part of a rectangular range); flagging them
    # caused widespread false-positive retries that destroyed correct
    # turn-1 answers. We only flag formula-string cells, which are an
    # objective bug (openpyxl never evaluates ``=`` strings, so the grader
    # reads None for them).
    formula_cells: list[str] = []
    sample_values: list[str] = []
    try:
        for scr in (answer_position or "").split(","):
            scr = scr.strip()
            if not scr:
                continue
            if "!" in scr:
                sheet_name, cell_range = scr.split("!", 1)
                sheet_name = sheet_name.strip().strip("'\"")
            else:
                sheet_name = wb.sheetnames[0] if wb.sheetnames else ""
                cell_range = scr
            cell_range = cell_range.strip().strip("'\"")
            if not sheet_name or sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            try:
                names = _generate_cell_names(cell_range)
            except Exception:
                continue
            for cn in names:
                try:
                    v = ws[cn].value
                except Exception:
                    continue
                ref = f"{sheet_name}!{cn}"
                if isinstance(v, str) and v.lstrip().startswith("="):
                    formula_cells.append(f"{ref}={v!r}")
                else:
                    if len(sample_values) < 8:
                        s = repr(v)
                        if len(s) > 60:
                            s = s[:57] + "..."
                        sample_values.append(f"{ref}={s}")
    finally:
        try:
            wb.close()
        except Exception:
            pass

    if not formula_cells:
        return None  # looks fine; don't re-prompt

    parts = ["## Self-check of answer-position cells in your output"]
    if formula_cells:
        parts.append(
            f"- {len(formula_cells)} target cell(s) contain a literal '='-formula "
            "string. openpyxl does NOT evaluate formulas, so the grader will see "
            "None for these. Compute the result in Python and write the literal "
            "value instead:"
        )
        for ref in formula_cells[:8]:
            parts.append(f"    {ref}")
        if len(formula_cells) > 8:
            parts.append(
                f"    ... and {len(formula_cells) - 8} more formula-string cells"
            )
    if sample_values:
        parts.append(
            "- Sample of other target cells in your output (these may still be "
            "wrong — double-check against the instruction):"
        )
        for s in sample_values[:8]:
            parts.append(f"    {s}")
    parts.append(
        "\nIf any of the above are wrong, return a corrected complete "
        "Python script inside a ```python``` block. If you believe the "
        "current output is actually correct as-is (e.g. the instruction "
        "explicitly says to leave cells empty), return your code unchanged."
    )
    return "\n".join(parts)


# ── Code extraction (same as official prompt.py) ────────────────────────────

def extract_code(text: str) -> str:
    """Extract the first ```python``` fenced code block from LLM output."""
    if "```" not in text:
        return text.strip()
    start = text.find("```")
    nl = text.find("\n", start)
    end = text.find("```", nl + 1)
    if nl == -1 or end == -1:
        return text.strip()
    return text[nl + 1 : end].strip()


# ── Prompt construction (official SpreadsheetBench prompts) ─────────────────


def _build_system(skill_content: str) -> str:
    base = load_prompt("codegen_system", env="spreadsheetbench")
    if skill_content.strip():
        base += f"\n\n## Skill\n{skill_content.strip()}"
    return base


def _build_user(
    instruction: str,
    input_xlsx: str,
    instruction_type: str = "",
    answer_position: str = "",
    diagnostic_mode: bool = False,
    diagnostic_instruction: str = "",
    diagnostic_trace_context: str = "",
) -> str:
    try:
        preview = _preview_workbook(input_xlsx)
    except Exception as e:  # noqa: BLE001
        preview = f"(failed to preview workbook: {e})"
    extra = ""
    if instruction_type:
        extra += f"\nInstruction type: {instruction_type}"
    if answer_position:
        extra += f"\nExpected answer position: {answer_position}"
    task_suffix = "Return only a ```python``` code block."
    diagnostic = ""
    if diagnostic_mode and diagnostic_instruction.strip():
        task_suffix = (
            "First provide a short diagnostic readout that follows the training "
            "instruction below, then return a single complete ```python``` code block."
        )
        diagnostic = f"\n\n# Training readout\n{diagnostic_instruction.strip()}"
    prefix = ""
    if diagnostic_trace_context.strip():
        prefix = (
            "# Previous Codex Trace Snapshot\n"
            "This is a partial transcript from an earlier attempt. Use it as your current reasoning context.\n\n"
            f"{diagnostic_trace_context.strip()}\n\n"
        )
    return (
        f"{prefix}"
        f"# Instruction\n{instruction}\n{extra}\n\n"
        f"# Input spreadsheet preview\n{preview}\n\n"
        "# Task\n"
        "Write a Python script that reads the workbook from the variable `INPUT_PATH`, "
        "applies the instruction, and writes the modified workbook to `OUTPUT_PATH`. "
        "Preserve all other cells unchanged. "
        "The preview may be truncated — do not hardcode row counts or assume the data ends at the last previewed row; "
        "iterate over all actual rows in the workbook instead. "
        f"{task_suffix}"
        f"{diagnostic}"
    )


# ── LLM call with retry ────────────────────────────────────────────────────

def _llm_call_with_retry(call_fn, *, retries: int = 5, timeout: int = 120):
    """Wrap an LLM API call with retry and per-call timeout."""
    last_err = None
    for attempt in range(retries):
        try:
            return call_fn(timeout=timeout)
        except Exception as e:  # noqa: BLE001
            last_err = e
            sleep = min(2 ** attempt + random.random(), 60)
            time.sleep(sleep)
    raise RuntimeError(f"LLM call failed after {retries} retries: {last_err}")


def _get_deployment() -> str:
    from skillopt.model import azure_openai as _llm
    return _llm.TARGET_DEPLOYMENT


def _build_codex_skill(skill_content: str) -> str:
    return render_skill_md(
        skill_content,
        description="Dynamic ReflACT skill for solving the current SpreadsheetBench task.",
        preamble=(
            "Use this skill when solving the current SpreadsheetBench task in this workspace.\n"
            "Write a single self-contained Python solution to `solution.py`.\n"
            "The solution must operate on the provided `INPUT_PATH` and `OUTPUT_PATH` variables.\n"
            "You may inspect `input.xlsx` and run `python run_solution.py` to validate locally,\n"
            "but do not hardcode values from the preview or from one specific workbook."
        ),
    )


def _build_codex_task(
    instruction: str,
    input_xlsx: str,
    instruction_type: str,
    answer_position: str,
    *,
    diagnostic_mode: bool,
    diagnostic_instruction: str,
    diagnostic_trace_context: str,
) -> str:
    prompt = _build_user(
        instruction,
        input_xlsx,
        instruction_type,
        answer_position,
        diagnostic_mode=diagnostic_mode,
        diagnostic_instruction=diagnostic_instruction,
        diagnostic_trace_context=diagnostic_trace_context,
    )
    return (
        f"{prompt}\n\n"
        "## Codex Harness Task\n"
        "- Read `.agents/skills/skillopt-target/SKILL.md` before writing code; do not call a Skill tool.\n"
        "- Read and optionally inspect `input.xlsx` in this workspace.\n"
        "- Write the final Python solution to `solution.py`.\n"
        "- The script should use the provided `INPUT_PATH` and `OUTPUT_PATH` variables.\n"
        "- If you want to validate locally, run `python run_solution.py`.\n"
        "- Do not return a code fence as the primary artifact; the source of truth is `solution.py`.\n"
    )


def _build_codex_driver() -> str:
    return (
        "import pathlib\n"
        "import re\n"
        "import sys\n"
        "import traceback\n\n"
        'INPUT_PATH = "input.xlsx"\n'
        'OUTPUT_PATH = "output.xlsx"\n'
        "code = pathlib.Path('solution.py').read_text(encoding='utf-8')\n"
        "code = re.sub(r'^\\s*(INPUT_PATH|OUTPUT_PATH)\\s*=\\s*.+$', '', code, flags=re.MULTILINE)\n"
        "globals_dict = {'__name__': '__main__', 'INPUT_PATH': INPUT_PATH, 'OUTPUT_PATH': OUTPUT_PATH}\n"
        "try:\n"
        "    exec(compile(code, 'solution.py', 'exec'), globals_dict, globals_dict)\n"
        "except Exception:\n"
        "    traceback.print_exc()\n"
        "    sys.exit(2)\n"
    )


def _prepare_codex_workspace(
    *,
    instruction: str,
    input_xlsx: str,
    output_path: str,
    instruction_type: str,
    answer_position: str,
    skill_content: str,
    diagnostic_mode: bool,
    diagnostic_instruction: str,
    diagnostic_trace_context: str,
    workspace_name: str = "codex_single",
) -> tuple[str, str, str, str]:
    task_out_dir = os.path.dirname(output_path)
    work_dir = os.path.join(task_out_dir, workspace_name)
    skill_md = _build_codex_skill(skill_content)
    task_md = _build_codex_task(
        instruction,
        input_xlsx,
        instruction_type,
        answer_position,
        diagnostic_mode=diagnostic_mode,
        diagnostic_instruction=diagnostic_instruction,
        diagnostic_trace_context=diagnostic_trace_context,
    )
    prompt = (
        "Read `.agents/skills/skillopt-target/SKILL.md` directly; do not call a Skill tool.\n"
        "Read `task.md`, inspect `input.xlsx` if useful, and write the final solution to `solution.py`.\n"
        "You may run `python run_solution.py` to validate the script locally.\n"
        "In your final response, briefly confirm whether `solution.py` was written and summarize the approach."
    )
    prepare_workspace(
        work_dir=work_dir,
        skill_md=skill_md,
        task_text=task_md,
        extra_files={"run_solution.py": _build_codex_driver()},
        copy_files=[(input_xlsx, "input.xlsx")],
    )

    return work_dir, skill_md, task_md, prompt


def _run_exec_backend(
    *,
    work_dir: str,
    prompt: str,
    model: str,
    timeout: int,
) -> tuple[str, str]:
    return run_target_exec(
        work_dir=work_dir,
        prompt=prompt,
        model=model,
        timeout=timeout,
        allow_file_edits=True,
    )


# ── Chat (no tools) ────────────────────────────────────────────────────────

def _chat_call(
    client,
    deployment: str,
    messages: list[dict],
    max_output_tokens: int,
    llm_timeout: int = 120,
    *,
    reasoning_effort_override: str | None = None,
    retries: int = 5,
) -> str:
    """Single LLM call, no tools. Returns raw text.

    Args:
        reasoning_effort_override: If not None, use this effort instead of
            the process-wide REASONING_EFFORT. Used by the timeout-fallback
            path to retry with a lower effort.
        retries: How many attempts _llm_call_with_retry makes before giving
            up (default 5, matching legacy behavior).
    """
    if reasoning_effort_override is not None:
        reasoning_effort = reasoning_effort_override
    else:
        reasoning_effort = get_reasoning_effort()
    if _needs_responses_api(deployment):
        # Responses API
        system = ""
        api_input = []
        for m in messages:
            if m["role"] == "system":
                system = m["content"]
            else:
                api_input.append({"role": m["role"], "content": m["content"]})
        resp = _llm_call_with_retry(lambda timeout: client.responses.create(
            model=deployment,
                instructions=system,
                input=api_input,
                max_output_tokens=max_output_tokens,
                **({"reasoning": {"effort": reasoning_effort}} if reasoning_effort else {}),
                timeout=timeout,
            ), timeout=llm_timeout, retries=retries)
        if hasattr(resp, "usage") and resp.usage:
            tracker.record(
                "rollout",
                getattr(resp.usage, "input_tokens", 0) or 0,
                getattr(resp.usage, "output_tokens", 0) or 0,
            )
        text = getattr(resp, "output_text", None) or ""
        if text:
            return text
        for item in getattr(resp, "output", None) or []:
            for part in getattr(item, "content", []):
                if getattr(part, "type", "") == "output_text":
                    return part.text or ""
        return ""
    else:
        # Chat Completions API — no tools
        kwargs = {
            "model": deployment,
            "messages": messages,
            "max_completion_tokens": max_output_tokens,
        }
        if reasoning_effort is not None:
            kwargs["reasoning_effort"] = reasoning_effort
        resp = _llm_call_with_retry(lambda timeout: client.chat.completions.create(
            **kwargs,
            timeout=timeout,
        ), timeout=llm_timeout, retries=retries)
        if resp.usage:
            tracker.record(
                "rollout",
                resp.usage.prompt_tokens or 0,
                resp.usage.completion_tokens or 0,
            )
        return resp.choices[0].message.content or ""


def _chat_call_with_fallback(
    client,
    deployment: str,
    messages: list[dict],
    max_output_tokens: int,
    *,
    llm_timeout: int,
    deadline: float,
    turn: int = 0,
) -> str:
    """Wrap `_chat_call` with a lower-effort retry on failure.

    Only active when SPREADSHEETBENCH_TIMEOUT_FALLBACK=1 (see module-level
    docstring). Cuts the first attempt's inner retries to 2 so we still
    have budget for a low-effort retry within the task deadline.

    On the second (fallback) attempt, ``reasoning_effort`` is overridden to
    `_FALLBACK_EFFORT` (default "low").
    """
    current_effort = get_reasoning_effort()
    if (
        not _TIMEOUT_FALLBACK
        or current_effort in (None, "low", "minimal")
        or current_effort == _FALLBACK_EFFORT
    ):
        # No fallback path available: caller already at low effort, or
        # the feature is off. Preserve original behaviour (retries=5).
        return _chat_call(
            client, deployment, messages, max_output_tokens,
            llm_timeout=llm_timeout,
        )

    # First attempt: fewer retries so we cap wasted time.
    try:
        return _chat_call(
            client, deployment, messages, max_output_tokens,
            llm_timeout=llm_timeout,
            retries=2,
        )
    except RuntimeError as e:
        remaining = int(deadline - time.time())
        if remaining <= 30:
            raise
        print(
            f"[timeout_fallback] turn={turn+1} effort={current_effort} failed "
            f"({str(e)[:180]}); retrying with effort={_FALLBACK_EFFORT}, "
            f"budget={remaining}s",
            flush=True,
        )
        return _chat_call(
            client, deployment, messages, max_output_tokens,
            llm_timeout=min(llm_timeout, remaining),
            reasoning_effort_override=_FALLBACK_EFFORT,
            retries=2,
        )


# ── Public API ──────────────────────────────────────────────────────────────

def run_single(
    instruction: str,
    input_xlsx: str,
    output_path: str,
    instruction_type: str = "",
    answer_position: str = "",
    skill_content: str = "",
    max_output_tokens: int = 16384,
    llm_timeout: int = 120,
    task_timeout: int = 300,
    diagnostic_mode: bool = False,
    diagnostic_instruction: str = "",
    diagnostic_trace_context: str = "",
) -> dict:
    """Single-round code generation. One LLM call, no tools.

    Args:
        llm_timeout: Per-LLM-call timeout in seconds (default 120).
        task_timeout: Total task timeout in seconds (default 300).

    Returns ``{"code": str, "raw": str, "n_turns": 1}``.
    """
    if is_target_exec_backend():
        deadline = time.time() + task_timeout
        deployment = _get_deployment()
        work_dir, skill_md, task_md, prompt = _prepare_codex_workspace(
            instruction=instruction,
            input_xlsx=input_xlsx,
            output_path=output_path,
            instruction_type=instruction_type,
            answer_position=answer_position,
            skill_content=skill_content,
            diagnostic_mode=diagnostic_mode,
            diagnostic_instruction=diagnostic_instruction,
            diagnostic_trace_context=diagnostic_trace_context,
        )
        remaining = max(10, int(deadline - time.time()))
        effective_timeout = min(task_timeout, remaining)
        final_message, raw = _run_exec_backend(
            work_dir=work_dir,
            prompt=prompt,
            model=deployment,
            timeout=effective_timeout,
        )
        solution_path = os.path.join(work_dir, "solution.py")
        if os.path.exists(solution_path):
            with open(solution_path, encoding="utf-8") as f:
                code = f.read()
        else:
            code = extract_code(final_message or raw)
        return {
            "code": code,
            "raw": raw or final_message,
            "n_turns": 1,
            "conversation": [{"role": "assistant", "content": final_message or raw}],
            "target_system_prompt": skill_md,
            "target_user_prompt": f"{prompt}\n\n## Task File\n\n{task_md}",
        }

    deadline = time.time() + task_timeout
    client = get_target_client()
    deployment = _get_deployment()
    system = _build_system(skill_content)
    user = _build_user(
        instruction,
        input_xlsx,
        instruction_type,
        answer_position,
        diagnostic_mode=diagnostic_mode,
        diagnostic_instruction=diagnostic_instruction,
        diagnostic_trace_context=diagnostic_trace_context,
    )

    messages = [
        {"role": "system", "content": system},
        {"role": "user", "content": user},
    ]

    remaining = max(10, int(deadline - time.time()))
    effective_timeout = min(llm_timeout, remaining)
    raw = _chat_call(client, deployment, messages, max_output_tokens, llm_timeout=effective_timeout)
    time.sleep(3)  # Rate-limit cooldown after successful LLM call
    code = extract_code(raw)

    return {
        "code": code,
        "raw": raw,
        "n_turns": 1,
        "conversation": [{"role": "assistant", "content": raw}],
        "target_system_prompt": system,
        "target_user_prompt": user,
    }


def run_multi(
    instruction: str,
    input_xlsx: str,
    output_path: str,
    instruction_type: str = "",
    answer_position: str = "",
    skill_content: str = "",
    max_turns: int = 5,
    max_output_tokens: int = 16384,
    llm_timeout: int = 120,
    task_timeout: int = 600,
    gold_path: str = "",
    diagnostic_mode: bool = False,
    diagnostic_instruction: str = "",
    diagnostic_trace_context: str = "",
) -> dict:
    """Multi-round code generation with execution feedback. No tools.

    Each round: LLM generates code → execute → if error, feed back and retry.

    Args:
        llm_timeout: Per-LLM-call timeout in seconds (default 120).
        task_timeout: Total task timeout in seconds (default 600).
        gold_path: Path to golden answer xlsx for eval feedback during
            training.  When non-empty, a successful execution is followed
            by an eval check; if the output is wrong the agent receives
            cell-level feedback (without revealing expected values) and
            gets another turn.  Leave empty for eval/test to avoid
            data leakage.

    Returns ``{"code": str, "raw": str, "n_turns": int, "conversation": [...]}``.
    """
    if is_target_exec_backend():
        deadline = time.time() + task_timeout
        deployment = _get_deployment()
        work_dir, skill_md, task_md, initial_prompt = _prepare_codex_workspace(
            instruction=instruction,
            input_xlsx=input_xlsx,
            output_path=output_path,
            instruction_type=instruction_type,
            answer_position=answer_position,
            skill_content=skill_content,
            diagnostic_mode=diagnostic_mode,
            diagnostic_instruction=diagnostic_instruction,
            diagnostic_trace_context=diagnostic_trace_context,
            workspace_name="codex_multi",
        )
        prompt = (
            f"{initial_prompt}\n\n"
            "## Multi-Turn Repair Mode\n"
            "- This is turn 1. Write or overwrite `solution.py`.\n"
            "- After each turn, the harness will execute your `solution.py`; if it fails, you will receive feedback and may revise it.\n"
            "- Keep the script general: use `INPUT_PATH` and `OUTPUT_PATH`, and do not hardcode one workbook's values."
        )
        conversation: list[dict] = []
        code = ""
        raw = ""
        final_message = ""
        solution_path = os.path.join(work_dir, "solution.py")

        for turn in range(max_turns):
            remaining = deadline - time.time()
            if remaining <= 10:
                break

            effective_timeout = max(10, int(remaining))
            final_message, raw = _run_exec_backend(
                work_dir=work_dir,
                prompt=prompt,
                model=deployment,
                timeout=effective_timeout,
            )
            conversation.append({"role": "assistant", "content": final_message or raw})

            if os.path.exists(solution_path):
                with open(solution_path, encoding="utf-8") as f:
                    code = f.read()
            else:
                code = extract_code(final_message or raw)
                if code.strip():
                    with open(solution_path, "w", encoding="utf-8") as f:
                        f.write(code)

            if not code.strip():
                feedback = (
                    "No usable `solution.py` or Python code block was produced. "
                    "Write a complete `solution.py` that reads `INPUT_PATH` and saves `OUTPUT_PATH`."
                )
            else:
                ok, err = run_generated_code(code, input_xlsx, output_path)
                if ok:
                    if gold_path and answer_position:
                        from .rollout import _auto_verify_output
                        eval_result = evaluate(
                            output_path, gold_path, instruction_type, answer_position,
                        )
                        if eval_result["ok"]:
                            break
                        verify = _auto_verify_output(output_path, gold_path, answer_position)
                        feedback = _build_eval_feedback(verify)
                    else:
                        # round-0 harness-opt: gold_path="" path used to break
                        # at turn 1, wasting the multi-loop. Self-introspect
                        # the answer-position cells; if structurally broken
                        # (None or literal '=...'), re-prompt once.
                        introspect = _self_introspect(output_path, answer_position)
                        if introspect is None:
                            break
                        feedback = (
                            "Your `solution.py` ran without errors, but a "
                            "structural sanity check of the answer-position "
                            "cells in your output suggests it may be incomplete.\n\n"
                            + introspect
                        )
                else:
                    feedback = (
                        "The current `solution.py` raised an error during harness execution:\n\n"
                        f"```\n{err[:3000]}\n```\n\n"
                        "Revise `solution.py` to fix the error. Keep using `INPUT_PATH` and `OUTPUT_PATH`."
                    )

            feedback_path = os.path.join(work_dir, f"feedback_turn_{turn + 1:02d}.md")
            with open(feedback_path, "w", encoding="utf-8") as f:
                f.write(feedback)
            conversation.append({"role": "user", "content": feedback})
            prompt = (
                f"The previous `solution.py` was evaluated and needs another revision.\n"
                f"Read `{os.path.basename(feedback_path)}` and update `solution.py` accordingly.\n"
                "You may run `python run_solution.py` for a local syntax/runtime check, but the harness will run the final code separately.\n"
                "Do not hardcode workbook-specific answers; preserve unrelated cells."
            )

        return {
            "code": code,
            "raw": raw or final_message,
            "n_turns": len([m for m in conversation if m["role"] == "assistant"]),
            "conversation": conversation,
            "target_system_prompt": skill_md,
            "target_user_prompt": f"{initial_prompt}\n\n## Task File\n\n{task_md}",
        }

    deadline = time.time() + task_timeout
    client = get_target_client()
    deployment = _get_deployment()
    system = _build_system(skill_content)
    user = _build_user(
        instruction,
        input_xlsx,
        instruction_type,
        answer_position,
        diagnostic_mode=diagnostic_mode,
        diagnostic_instruction=diagnostic_instruction,
        diagnostic_trace_context=diagnostic_trace_context,
    )

    messages: list[dict] = [
        {"role": "system", "content": system},
        {"role": "user", "content": user},
    ]
    conversation: list[dict] = []
    code = ""
    raw = ""

    for turn in range(max_turns):
        remaining = deadline - time.time()
        if remaining <= 10:
            # Not enough time for another round
            break

        effective_timeout = min(llm_timeout, int(remaining))
        raw = _chat_call_with_fallback(
            client, deployment, messages, max_output_tokens,
            llm_timeout=effective_timeout,
            deadline=deadline,
            turn=turn,
        )
        time.sleep(3)  # Rate-limit cooldown after successful LLM call
        code = extract_code(raw)
        conversation.append({"role": "assistant", "content": raw})
        messages.append({"role": "assistant", "content": raw})

        if not code.strip():
            # No code extracted — ask again
            feedback = (
                "No Python code block was found in your response. "
                "Please return a complete Python script inside a ```python``` block."
            )
            messages.append({"role": "user", "content": feedback})
            conversation.append({"role": "user", "content": feedback})
            continue

        # Execute the code
        ok, err = run_generated_code(code, input_xlsx, output_path)
        if ok:
            # Execution succeeded — check correctness if gold_path available
            if gold_path and answer_position:
                from .rollout import _auto_verify_output
                eval_result = evaluate(
                    output_path, gold_path, instruction_type, answer_position,
                )
                if eval_result["ok"]:
                    break  # Genuinely correct — stop

                # Output is wrong — build feedback without leaking golden values
                verify = _auto_verify_output(output_path, gold_path, answer_position)
                feedback = _build_eval_feedback(verify)
                messages.append({"role": "user", "content": feedback})
                conversation.append({"role": "user", "content": feedback})
                continue
            else:
                # No gold path (eval/test). Run gold-free self-introspect on
                # the answer-position cells; if structurally broken (None at
                # target / literal '=...' string), re-prompt once. Otherwise
                # accept execution success and stop.
                introspect = _self_introspect(output_path, answer_position)
                if introspect is None:
                    break
                feedback = (
                    "Your code ran without errors, but a structural sanity "
                    "check of the answer-position cells in your output "
                    "suggests it may be incomplete.\n\n" + introspect
                )
                messages.append({"role": "user", "content": feedback})
                conversation.append({"role": "user", "content": feedback})
                continue

        # Execution failed — feed error back
        feedback = (
            f"The code raised an error during execution:\n\n"
            f"```\n{err[:3000]}\n```\n\n"
            f"Please fix the code and return a complete corrected Python script "
            f"inside a ```python``` block."
        )
        messages.append({"role": "user", "content": feedback})
        conversation.append({"role": "user", "content": feedback})

    return {
        "code": code,
        "raw": raw,
        "n_turns": turn + 1,
        "conversation": conversation,
        "target_system_prompt": system,
        "target_user_prompt": user,
    }
