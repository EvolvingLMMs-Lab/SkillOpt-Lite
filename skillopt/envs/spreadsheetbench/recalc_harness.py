"""Optional, opt-in recalc + verify harness for SpreadsheetBench.

This module is **separate from the default eval path** and is **disabled by
default**. It exists to close the single biggest residual failure cluster:
the agent writes an Excel formula string (e.g. ``=SUMPRODUCT(...)``) into a
cell, but ``openpyxl`` never evaluates formulas, so the grader reading with
``data_only=True`` sees ``None``.

When enabled, it adds:

1. A ``recalc_xlsx`` tool the ReAct agent can call to materialize formula
   results into cached cell values (so the grader sees numbers, not ``None``),
   and to read those values back.
2. An optional post-exec auto-recalc note appended after ``solution.py`` runs.

Nothing here runs unless one of the env vars below is set, so the original
harness behaves identically when the feature is off.

Enable flags (all default OFF):
    SPREADSHEETBENCH_RECALC=1        -> expose the ``recalc_xlsx`` agent tool
    SPREADSHEETBENCH_RECALC_AUTO=1   -> also auto-recalc the output after each
                                        ``solution.py`` run (implies RECALC)

Backends (auto-detected, tried in order):
    - LibreOffice headless (``soffice`` / ``libreoffice``)
    - the ``formulas`` Python library
Neither is required for the default path. If the feature is enabled but no
backend is installed, the tool returns a clear, non-fatal message.
"""
from __future__ import annotations

import os
import shutil
import subprocess
import tempfile


# ── Feature flags ─────────────────────────────────────────────────────────────

_TRUTHY = {"1", "true", "yes", "on"}


def _flag(name: str) -> bool:
    return str(os.environ.get(name, "")).strip().lower() in _TRUTHY


def recalc_enabled() -> bool:
    """True if the agent-facing recalc tool should be exposed."""
    return _flag("SPREADSHEETBENCH_RECALC") or recalc_auto_enabled()


def recalc_auto_enabled() -> bool:
    """True if the output should be auto-recalced after solution.py runs."""
    return _flag("SPREADSHEETBENCH_RECALC_AUTO")


# ── Backend detection ─────────────────────────────────────────────────────────

def _soffice_bin() -> str | None:
    for name in ("soffice", "libreoffice"):
        path = shutil.which(name)
        if path:
            return path
    return None


def _has_formulas_lib() -> bool:
    try:  # pragma: no cover - import probe
        import importlib.util

        return importlib.util.find_spec("formulas") is not None
    except Exception:  # noqa: BLE001
        return False


def available_backend() -> str | None:
    """Return the name of the first available recalc backend, else None."""
    if _soffice_bin():
        return "libreoffice"
    if _has_formulas_lib():
        return "formulas"
    return None


# ── Recalc backends ───────────────────────────────────────────────────────────

def _recalc_libreoffice(path: str, timeout: int) -> tuple[bool, str]:
    bin_path = _soffice_bin()
    if not bin_path:
        return False, "libreoffice not found"
    tmp_out = tempfile.mkdtemp(prefix="ssb_recalc_")
    tmp_profile = tempfile.mkdtemp(prefix="ssb_loprofile_")
    try:
        proc = subprocess.run(
            [
                bin_path,
                "--headless",
                "--calc",
                f"-env:UserInstallation=file://{tmp_profile}",
                "--convert-to",
                "xlsx:Calc MS Excel 2007 XML",
                "--outdir",
                tmp_out,
                path,
            ],
            capture_output=True,
            text=True,
            timeout=timeout,
        )
        base = os.path.splitext(os.path.basename(path))[0] + ".xlsx"
        produced = os.path.join(tmp_out, base)
        if proc.returncode != 0 or not os.path.exists(produced):
            return False, (proc.stdout + "\n" + proc.stderr).strip()[:500]
        shutil.copyfile(produced, path)
        return True, "recalced via libreoffice"
    except subprocess.TimeoutExpired:
        return False, f"libreoffice recalc timeout after {timeout}s"
    except Exception as e:  # noqa: BLE001
        return False, f"libreoffice recalc error: {e}"
    finally:
        shutil.rmtree(tmp_out, ignore_errors=True)
        shutil.rmtree(tmp_profile, ignore_errors=True)


def _recalc_formulas_lib(path: str, timeout: int) -> tuple[bool, str]:
    """Recalc via the ``formulas`` lib, patching values into the ORIGINAL file.

    The lib's own ``.write()`` rebuilds the workbook and mangles sheet names /
    drops formatting (causing "worksheet not found" regressions on sheet_level
    tasks). Instead we compute the solution dict, then overwrite only the
    formula cells' values in the original openpyxl workbook -- preserving sheet
    names, merged cells, and all formatting.
    """
    try:
        import re

        import formulas  # type: ignore
        import openpyxl
    except Exception as e:  # noqa: BLE001
        return False, f"formulas lib import failed: {e}"
    try:
        sol = formulas.ExcelModel().loads(path).finish(circular=True).calculate()
    except Exception as e:  # noqa: BLE001
        return False, f"formulas lib calc error: {e}"

    # Keys look like "'[book.xlsx]SHEET NAME'!A1" (sheet names are uppercased).
    key_re = re.compile(r"^'?\[[^\]]*\]([^']*)'?!\$?([A-Za-z]+)\$?(\d+)$")
    val_map: dict[tuple[str, str], object] = {}
    for k, v in sol.items():
        m = key_re.match(str(k))
        if not m:
            continue
        sheet = m.group(1).upper()
        coord = (m.group(2) + m.group(3)).upper()
        try:
            scalar = v.value[0, 0]
        except Exception:  # noqa: BLE001
            try:
                scalar = v.value
            except Exception:  # noqa: BLE001
                continue
        val_map[(sheet, coord)] = _coerce(scalar)

    if not val_map:
        return False, "formulas lib produced no values"

    try:
        wb = openpyxl.load_workbook(path, data_only=False)
    except Exception as e:  # noqa: BLE001
        return False, f"reopen failed: {e}"
    patched = 0
    try:
        for ws in wb.worksheets:
            sheet_up = ws.title.upper()
            for row in ws.iter_rows():
                for c in row:
                    if isinstance(c.value, str) and c.value.startswith("="):
                        nv = val_map.get((sheet_up, c.coordinate.upper()))
                        if nv is not None:
                            c.value = nv
                            patched += 1
        wb.save(path)
    except Exception as e:  # noqa: BLE001
        return False, f"patch/save failed: {e}"
    finally:
        wb.close()
    return True, (
        f"recalced via formulas lib (patched {patched} cells, structure preserved)"
    )


def _coerce(v: object) -> object:
    """Convert numpy/odd scalar types to plain Python so openpyxl can store them."""
    try:
        import numpy as np  # type: ignore

        if isinstance(v, np.generic):
            return v.item()
    except Exception:  # noqa: BLE001
        pass
    return v


def recalc_workbook(path: str, timeout: int = 120) -> tuple[bool, str]:
    """Recalculate all formulas in ``path`` in-place, caching their values.

    Returns ``(ok, message)``. Always non-fatal: on any failure returns
    ``(False, reason)`` and leaves the file untouched.
    """
    if not os.path.exists(path):
        return False, f"file not found: {path}"
    backend = available_backend()
    if backend == "libreoffice":
        return _recalc_libreoffice(path, timeout)
    if backend == "formulas":
        return _recalc_formulas_lib(path, timeout)
    return (
        False,
        "no recalc backend available (install LibreOffice or `pip install formulas`)",
    )


# ── Read-back inspection ──────────────────────────────────────────────────────

def _needs_recalc(path: str) -> bool:
    """True if the file has >=1 formula cell whose cached value is None.

    Used to bound the blast radius of auto-recalc: files that already have
    concrete values are left untouched, so a currently-passing output is never
    rewritten (the `formulas` backend rebuilds the workbook and can drop
    formatting). Only outputs exhibiting the exact failure signature -- a
    formula whose result the grader would read as None -- are recalced.
    """
    try:
        import openpyxl

        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception:  # noqa: BLE001
        return False
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    if c.value is None and c.data_type == "f":
                        return True
    finally:
        wb.close()
    # Fallback: formula text present but no cached value.
    try:
        wb_f = openpyxl.load_workbook(path, data_only=False)
        wb_v = openpyxl.load_workbook(path, data_only=True)
    except Exception:  # noqa: BLE001
        return False
    try:
        for sn in wb_f.sheetnames:
            ws_f, ws_v = wb_f[sn], wb_v[sn]
            for row_f, row_v in zip(ws_f.iter_rows(), ws_v.iter_rows()):
                for cf, cv in zip(row_f, row_v):
                    if (
                        isinstance(cf.value, str)
                        and cf.value.startswith("=")
                        and cv.value is None
                    ):
                        return True
    finally:
        wb_f.close()
        wb_v.close()
    return False


def _readback_summary(path: str, max_cells: int = 12) -> str:
    """Report cached values for cells that contain formulas (post-recalc)."""
    try:
        import openpyxl
    except Exception as e:  # noqa: BLE001
        return f"[readback error: {e}]"
    try:
        wb_f = openpyxl.load_workbook(path, data_only=False)
        wb_v = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:  # noqa: BLE001
        return f"[readback error: {e}]"
    lines: list[str] = []
    still_none: list[str] = []
    shown = 0
    try:
        for sn in wb_f.sheetnames:
            ws_f = wb_f[sn]
            ws_v = wb_v[sn]
            for row_f, row_v in zip(ws_f.iter_rows(), ws_v.iter_rows()):
                for cf, cv in zip(row_f, row_v):
                    fv = cf.value
                    if isinstance(fv, str) and fv.startswith("="):
                        if cv.value is None:
                            still_none.append(f"{sn}!{cf.coordinate}")
                        elif shown < max_cells:
                            lines.append(
                                f"    {sn}!{cf.coordinate}: {fv} = {cv.value!r}"
                            )
                            shown += 1
    finally:
        wb_f.close()
        wb_v.close()
    out = []
    if lines:
        out.append("  formula cells now cached:")
        out.extend(lines)
    if still_none:
        out.append(
            f"  WARNING: {len(still_none)} formula cells STILL have no cached "
            f"value: {still_none[:10]}"
        )
    if not out:
        # Some backends (e.g. the `formulas` lib) replace formulas with literal
        # results, so no `=` cells remain. That is expected and correct: the
        # grader reads concrete values. Inspect the file directly to confirm.
        out.append(
            "  values materialized (formulas replaced by literal results); "
            "read the target cells to confirm"
        )
    return "\n".join(out)


# ── Agent-facing tool ─────────────────────────────────────────────────────────

RECALC_TOOL_CHAT = {
    "type": "function",
    "function": {
        "name": "recalc_xlsx",
        "description": (
            "Recalculate all Excel formulas in an .xlsx file in-place so their "
            "results are cached as concrete cell values (the grader reads cached "
            "values; openpyxl never evaluates formulas). Use this after writing "
            "real Excel formulas, then read the cached values back. Returns the "
            "computed values for formula cells."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "path": {
                    "type": "string",
                    "description": "Path to the .xlsx file to recalculate in-place.",
                }
            },
            "required": ["path"],
        },
    },
}


def extra_tools_chat() -> list[dict]:
    """Tool schemas to add to the ReAct loop (empty unless enabled)."""
    return [RECALC_TOOL_CHAT] if recalc_enabled() else []


def run_recalc_tool(path: str, work_dir: str) -> str:
    """Dispatch handler for the ``recalc_xlsx`` tool."""
    full = path if os.path.isabs(path) else os.path.join(work_dir, path)
    if not os.path.exists(full):
        return f"[recalc_xlsx] file not found: {full}"
    ok, msg = recalc_workbook(full)
    if not ok:
        return f"[recalc_xlsx] recalc failed ({msg}). File left unchanged."
    return f"[recalc_xlsx] {msg}\n" + _readback_summary(full)


# ── Optional post-exec auto-recalc note ───────────────────────────────────────

def discover_output_path(work_dir: str) -> str | None:
    """Best-effort discovery of the output xlsx a solution.py wrote.

    Mirrors the lookup the existing auto-verify uses: prefer OUTPUT_PATH from
    solution.py, else the first ``*_pred*.xlsx`` in the work dir.
    """
    import glob as _glob

    sol_path = os.path.join(work_dir, "solution.py")
    if os.path.exists(sol_path):
        try:
            with open(sol_path) as f:
                for line in f:
                    stripped = line.strip()
                    if stripped.startswith("OUTPUT_PATH"):
                        val = stripped.split("=", 1)[1].strip().strip("'\"").strip()
                        if val:
                            return val
                        break
        except Exception:  # noqa: BLE001
            pass
    candidates = [
        f for f in _glob.glob(os.path.join(work_dir, "*.xlsx"))
        if "_pred" in os.path.basename(f)
    ]
    return candidates[0] if candidates else None


def maybe_recalc_path(path: str | None) -> str:
    """If auto-recalc is enabled, recalc the given output file in-place.

    Agent-agnostic hook intended to run right before grading. Returns a short
    note string (empty when the feature is off) so callers can append
    unconditionally without changing default behavior.
    """
    if not recalc_auto_enabled():
        return ""
    if not path or not os.path.exists(path):
        return ""
    # Only touch outputs that exhibit the failure signature, so currently
    # passing files are never rewritten (avoids formatting-loss regressions).
    if not _needs_recalc(path):
        return ""
    ok, msg = recalc_workbook(path)
    if not ok:
        return f"[recalc] skipped: {msg}"
    return f"[recalc] {msg}"


def maybe_post_exec_note(work_dir: str) -> str:
    """If auto-recalc is enabled, recalc the solution output and return a note.

    Returns an empty string when the feature is off so the caller can append
    unconditionally without changing default behavior.
    """
    if not recalc_auto_enabled():
        return ""
    output_path = discover_output_path(work_dir)
    if not output_path or not os.path.exists(output_path):
        return ""
    if not _needs_recalc(output_path):
        return ""
    ok, msg = recalc_workbook(output_path)
    if not ok:
        return f"\n\n[AUTO-RECALC] skipped: {msg}"
    return f"\n\n[AUTO-RECALC] {msg}\n" + _readback_summary(output_path)
